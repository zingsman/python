import openpyxl
import datetime,time
import random
def work(sheet):
    avg_column = sheet.max_column + 1
    sum_column = sheet.max_column + 2
    sheet.cell(row=1,column=avg_column).value = '平均分'
    sheet.cell(row=1,column=sum_column).value = '总分'
    for row in sheet.iter_rows(min_row=2,min_col=2,max_col=4):
        scores = [i.value for i in row]
        sum_scores = sum(scores)
        avg_scores = round(sum_scores / len(scores),2)
        sheet.cell(row=row[0].row,column=avg_column).value = avg_scores
        sheet.cell(row=row[0].row,column=sum_column).value = sum_scores
def main():
    #filename = 'C:\\Users\\ke.peng\\Desktop\\test\\pytest' + time.time() + '.xlsx'
    filename = str(random.randrange(1,100))
    wb=openpyxl.load_workbook('C:\\Users\\ke.peng\\Desktop\\test\\pytest.xlsx')
    ws=wb.get_sheet_by_name('一年级')
    work(ws)
    #wb.save('C:\\Users\\ke.peng\\Desktop\\test\\pytest2.xlsx')
    wb.save('C:\\Users\\ke.peng\\Desktop\\test\\pytest' + filename + '.xlsx')
    wb.save(filename)
if __name__ =='__main__':
    main()